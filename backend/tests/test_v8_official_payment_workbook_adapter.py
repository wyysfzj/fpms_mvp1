from __future__ import annotations

from hashlib import sha256
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from app.modules.annuity import verified_official_payment_workbook as adapter
from app.modules.annuity.verified_official_payment_workbook import (
    EXPECTED_HEADERS,
    EXPECTED_SHEETS,
    EXPECTED_STATES,
    EXPECTED_VALIDATIONS,
    VBA_PART,
    InvalidOfficialPaymentWorkbookError,
    OfficialPaymentRow,
    fill_template,
    validate_template,
)

FIXTURE = Path(__file__).parent / "fixtures" / "v8_verified_official_payment_template.xlsm"
SYNTHETIC_VBA = b"TEST_ONLY_SYNTHETIC_VBA_PROJECT_NO_EXECUTABLE_CODE\x00"


def _vba_bytes(source: Path | BytesIO) -> bytes:
    with ZipFile(source) as package:
        return package.read(VBA_PART)


def _rewrite_package(source: Path, destination: Path, replacements: dict[str, bytes]) -> None:
    with ZipFile(source) as original, ZipFile(destination, "w") as rewritten:
        for member in original.infolist():
            rewritten.writestr(member, replacements.get(member.filename, original.read(member)))


def test_validate_template_returns_exact_test_only_structure_snapshot() -> None:
    assert _vba_bytes(FIXTURE) == SYNTHETIC_VBA
    snapshot = validate_template(FIXTURE)

    assert snapshot.sheet_names == EXPECTED_SHEETS
    assert snapshot.sheet_states == EXPECTED_STATES
    assert snapshot.headers == EXPECTED_HEADERS
    assert snapshot.data_validations == EXPECTED_VALIDATIONS
    assert len(snapshot.column_widths) == 9
    assert all(width > 0 for width in snapshot.column_widths)
    assert snapshot.vba_project_sha256 == sha256(SYNTHETIC_VBA).hexdigest()


def test_fill_template_preserves_package_structure_and_writes_only_upload_rows(
    tmp_path: Path,
) -> None:
    fixture_before = FIXTURE.read_bytes()
    before = validate_template(FIXTURE)
    rows = (
        OfficialPaymentRow(
            sequence_number=1,
            application_number="TEST-APPLICATION-001",
            business_type="专利",
            invoice_title="测试机构",
            unified_social_credit_code="TEST_ONLY_CREDIT_CODE",
            fee_type="申请费",
            foreign_currency_amount=None,
            amount_cny=900,
            remark="TEST_ONLY",
        ),
        OfficialPaymentRow(
            sequence_number=2,
            application_number="TEST-APPLICATION-002",
            business_type="专利",
            invoice_title="测试机构",
            unified_social_credit_code="TEST_ONLY_CREDIT_CODE",
            fee_type="年费",
            foreign_currency_amount=10,
            amount_cny=600,
        ),
    )

    rendered = fill_template(FIXTURE, rows)
    rendered_path = tmp_path / "rendered.xlsm"
    rendered_path.write_bytes(rendered)

    assert FIXTURE.read_bytes() == fixture_before
    assert validate_template(rendered_path) == before
    assert _vba_bytes(BytesIO(rendered)) == SYNTHETIC_VBA

    workbook = load_workbook(BytesIO(rendered), keep_vba=True, data_only=False)
    upload = workbook[EXPECTED_SHEETS[0]]
    assert tuple(upload.cell(2, column).value for column in range(1, 10)) == rows[0].as_cells()
    assert tuple(upload.cell(3, column).value for column in range(1, 10)) == rows[1].as_cells()
    assert tuple(workbook[name].sheet_state for name in workbook.sheetnames) == EXPECTED_STATES


def test_invalid_structure_and_excess_rows_fail_before_output(tmp_path: Path) -> None:
    malformed_path = tmp_path / "malformed.xlsm"
    workbook = load_workbook(FIXTURE, keep_vba=True, data_only=False)
    workbook[EXPECTED_SHEETS[0]].cell(1, 1, "错误标题")
    workbook.save(malformed_path)

    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="headers"):
        fill_template(malformed_path, ())

    row = OfficialPaymentRow(1, "TEST", "专利", "测试", "TEST", "申请费", None, 1)
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="at most 500"):
        fill_template(FIXTURE, (row,) * 501)


def test_missing_malformed_and_misbound_packages_fail_with_domain_error(
    tmp_path: Path,
) -> None:
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="cannot be accessed"):
        validate_template(tmp_path / "missing.xlsm")

    malformed = tmp_path / "malformed.xlsm"
    malformed.write_bytes(b"not-an-ooxml-package")
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="readable OOXML"):
        validate_template(malformed)

    misbound = tmp_path / "misbound.xlsm"
    with ZipFile(FIXTURE) as package:
        relationships = package.read("xl/_rels/workbook.xml.rels").replace(
            b'Target="vbaProject.bin"', b'Target="otherProject.bin"'
        )
    _rewrite_package(
        FIXTURE,
        misbound,
        {"xl/_rels/workbook.xml.rels": relationships},
    )
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="declared VBA"):
        validate_template(misbound)

    wrong_content_type = tmp_path / "wrong-content-type.xlsm"
    with ZipFile(FIXTURE) as package:
        content_types = package.read("[Content_Types].xml").replace(
            b"application/vnd.ms-office.vbaProject",
            b"application/octet-stream",
        )
    _rewrite_package(
        FIXTURE,
        wrong_content_type,
        {"[Content_Types].xml": content_types},
    )
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="declared VBA"):
        validate_template(wrong_content_type)


def test_package_resource_limits_apply_before_openpyxl_parse(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(adapter, "MAX_PACKAGE_BYTES", FIXTURE.stat().st_size - 1)
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="size limit"):
        validate_template(FIXTURE)

    monkeypatch.setattr(adapter, "MAX_PACKAGE_BYTES", 25 * 1024 * 1024)
    monkeypatch.setattr(adapter, "MAX_PACKAGE_MEMBERS", 1)
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="too many members"):
        validate_template(FIXTURE)

    high_ratio = tmp_path / "high-ratio.xlsm"
    _rewrite_package(FIXTURE, high_ratio, {})
    monkeypatch.setattr(adapter, "MAX_PACKAGE_MEMBERS", 512)
    monkeypatch.setattr(adapter, "MAX_MEMBER_UNCOMPRESSED_BYTES", 1)
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="member exceeds"):
        validate_template(FIXTURE)

    monkeypatch.setattr(adapter, "MAX_MEMBER_UNCOMPRESSED_BYTES", 20 * 1024 * 1024)
    monkeypatch.setattr(adapter, "MAX_TOTAL_UNCOMPRESSED_BYTES", 1)
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="uncompressed size"):
        validate_template(FIXTURE)

    monkeypatch.setattr(adapter, "MAX_TOTAL_UNCOMPRESSED_BYTES", 100 * 1024 * 1024)
    monkeypatch.setattr(adapter, "MAX_COMPRESSION_RATIO", 0.5)
    with pytest.raises(InvalidOfficialPaymentWorkbookError, match="compression-ratio"):
        validate_template(high_ratio)
