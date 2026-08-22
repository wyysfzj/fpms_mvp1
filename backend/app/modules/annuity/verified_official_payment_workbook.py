from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from posixpath import normpath
from typing import Sequence
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

UPLOAD_SHEET = "网上缴费"
EXPECTED_SHEETS = (UPLOAD_SHEET, "Sheet2", "sheet1")
EXPECTED_STATES = ("visible", "hidden", "hidden")
EXPECTED_HEADERS = (
    "序号",
    "申请号/专利号/国际申请号/海牙转交编号",
    "业务类型",
    "票据抬头",
    "统一社会信用代码",
    "费用种类",
    "外币金额",
    "费用金额（人民币）",
    "备注",
)
VBA_PART = "xl/vbaProject.bin"
MAX_PACKAGE_BYTES = 25 * 1024 * 1024
MAX_PACKAGE_MEMBERS = 512
MAX_MEMBER_UNCOMPRESSED_BYTES = 20 * 1024 * 1024
MAX_TOTAL_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_COMPRESSION_RATIO = 100

CONTENT_TYPES_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/content-types"
RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
VBA_RELATIONSHIP_TYPE = "http://schemas.microsoft.com/office/2006/relationships/vbaProject"
VBA_CONTENT_TYPE = "application/vnd.ms-office.vbaProject"


class InvalidOfficialPaymentWorkbookError(ValueError):
    """Raised when a workbook does not match the frozen upload structure."""


@dataclass(frozen=True)
class DataValidationSnapshot:
    validation_type: str | None
    formula1: str | None
    operator: str | None
    allow_blank: bool
    ranges: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookStructureSnapshot:
    sheet_names: tuple[str, ...]
    sheet_states: tuple[str, ...]
    headers: tuple[str | None, ...]
    column_widths: tuple[float, ...]
    data_validations: tuple[DataValidationSnapshot, ...]
    vba_project_sha256: str


@dataclass(frozen=True)
class OfficialPaymentRow:
    sequence_number: int
    application_number: str
    business_type: str
    invoice_title: str
    unified_social_credit_code: str
    fee_type: str
    foreign_currency_amount: int | float | None
    amount_cny: int | float
    remark: str | None = None

    def as_cells(self) -> tuple[str | int | float | None, ...]:
        return (
            self.sequence_number,
            self.application_number,
            self.business_type,
            self.invoice_title,
            self.unified_social_credit_code,
            self.fee_type,
            self.foreign_currency_amount,
            self.amount_cny,
            self.remark,
        )


EXPECTED_VALIDATIONS = (
    DataValidationSnapshot(
        validation_type="list",
        formula1="'Sheet2'!$A$2:$A$3",
        operator=None,
        allow_blank=False,
        ranges=("C2:C501",),
    ),
    DataValidationSnapshot(
        validation_type="list",
        formula1="'sheet1'!$A$2:$A$3",
        operator=None,
        allow_blank=False,
        ranges=("F2:F501",),
    ),
    DataValidationSnapshot(
        validation_type="decimal",
        formula1="0",
        operator="greaterThanOrEqual",
        allow_blank=True,
        ranges=("G2:H501",),
    ),
)


def _validation_snapshot(validation: DataValidation) -> DataValidationSnapshot:
    return DataValidationSnapshot(
        validation_type=validation.type,
        formula1=validation.formula1,
        operator=validation.operator,
        allow_blank=bool(validation.allow_blank),
        ranges=tuple(str(cell_range) for cell_range in validation.sqref.ranges),
    )


def _source_size(source: Path | BytesIO) -> int:
    try:
        if isinstance(source, Path):
            return source.stat().st_size
        return source.getbuffer().nbytes
    except OSError as exc:
        raise InvalidOfficialPaymentWorkbookError("workbook package cannot be accessed") from exc


def _validated_package(source: Path | BytesIO) -> tuple[bytes, bytes, bytes]:
    if _source_size(source) > MAX_PACKAGE_BYTES:
        raise InvalidOfficialPaymentWorkbookError("workbook package exceeds the size limit")
    if isinstance(source, BytesIO):
        source.seek(0)
    try:
        with ZipFile(source) as package:
            members = package.infolist()
            if len(members) > MAX_PACKAGE_MEMBERS:
                raise InvalidOfficialPaymentWorkbookError(
                    "workbook package contains too many members"
                )
            if len({member.filename for member in members}) != len(members):
                raise InvalidOfficialPaymentWorkbookError(
                    "workbook package contains duplicate member names"
                )
            total_uncompressed = 0
            for member in members:
                if member.flag_bits & 0x1:
                    raise InvalidOfficialPaymentWorkbookError(
                        "encrypted workbook package members are not supported"
                    )
                if member.file_size > MAX_MEMBER_UNCOMPRESSED_BYTES:
                    raise InvalidOfficialPaymentWorkbookError(
                        "workbook package member exceeds the size limit"
                    )
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_TOTAL_UNCOMPRESSED_BYTES:
                    raise InvalidOfficialPaymentWorkbookError(
                        "workbook package exceeds the uncompressed size limit"
                    )
                if (member.file_size > 0 and member.compress_size == 0) or (
                    member.compress_size > 0
                    and member.file_size / member.compress_size > MAX_COMPRESSION_RATIO
                ):
                    raise InvalidOfficialPaymentWorkbookError(
                        "workbook package member exceeds the compression-ratio limit"
                    )
            project = package.read(VBA_PART)
            content_types = package.read("[Content_Types].xml")
            workbook_relationships = package.read("xl/_rels/workbook.xml.rels")
    except InvalidOfficialPaymentWorkbookError:
        raise
    except (BadZipFile, KeyError, OSError) as exc:
        raise InvalidOfficialPaymentWorkbookError(
            "workbook must be a readable OOXML package with a VBA project part"
        ) from exc
    return project, content_types, workbook_relationships


def _vba_project_bytes(source: Path | BytesIO) -> bytes:
    project, content_types, workbook_relationships = _validated_package(source)

    try:
        content_root = ElementTree.fromstring(content_types)
        relationships_root = ElementTree.fromstring(workbook_relationships)
    except ElementTree.ParseError as exc:
        raise InvalidOfficialPaymentWorkbookError(
            "workbook package relationship metadata is malformed"
        ) from exc

    relationships = relationships_root.findall(f"{{{RELATIONSHIPS_NAMESPACE}}}Relationship")

    def resolved_target(relationship: ElementTree.Element) -> str:
        target = relationship.get("Target", "")
        return normpath(target if target.startswith("/") else f"/xl/{target}")

    vba_relationships = [
        relationship
        for relationship in relationships
        if relationship.get("Type") == VBA_RELATIONSHIP_TYPE
        and relationship.get("TargetMode") != "External"
        and resolved_target(relationship) == f"/{VBA_PART}"
    ]
    override_matches = any(
        override.get("PartName") == f"/{VBA_PART}"
        and override.get("ContentType") == VBA_CONTENT_TYPE
        for override in content_root.findall(f"{{{CONTENT_TYPES_NAMESPACE}}}Override")
    )
    default_matches = any(
        default.get("Extension") == "bin" and default.get("ContentType") == VBA_CONTENT_TYPE
        for default in content_root.findall(f"{{{CONTENT_TYPES_NAMESPACE}}}Default")
    )

    if not project or len(vba_relationships) != 1 or not (override_matches or default_matches):
        raise InvalidOfficialPaymentWorkbookError(
            "workbook must contain a declared VBA project package part"
        )
    return project


def _structure_snapshot(source: Path | BytesIO) -> WorkbookStructureSnapshot:
    vba_project = _vba_project_bytes(source)
    if isinstance(source, BytesIO):
        source.seek(0)
    try:
        workbook = load_workbook(source, keep_vba=True, data_only=False)
    except Exception as exc:
        raise InvalidOfficialPaymentWorkbookError(
            "workbook cannot be opened as an .xlsm template"
        ) from exc

    sheet_names = tuple(workbook.sheetnames)
    sheet_states = tuple(sheet.sheet_state for sheet in workbook.worksheets)
    if sheet_names != EXPECTED_SHEETS or sheet_states != EXPECTED_STATES:
        raise InvalidOfficialPaymentWorkbookError(
            "workbook sheet order or visibility does not match the expected template"
        )

    upload_sheet = workbook[UPLOAD_SHEET]
    headers = tuple(upload_sheet.cell(row=1, column=index).value for index in range(1, 10))
    if headers != EXPECTED_HEADERS:
        raise InvalidOfficialPaymentWorkbookError(
            "upload sheet headers do not match the expected template"
        )

    validations = tuple(
        _validation_snapshot(validation)
        for validation in upload_sheet.data_validations.dataValidation
    )
    if validations != EXPECTED_VALIDATIONS:
        raise InvalidOfficialPaymentWorkbookError(
            "upload sheet data validations do not match the expected template"
        )

    widths = tuple(float(upload_sheet.column_dimensions[letter].width) for letter in "ABCDEFGHI")
    return WorkbookStructureSnapshot(
        sheet_names=sheet_names,
        sheet_states=sheet_states,
        headers=headers,
        column_widths=widths,
        data_validations=validations,
        vba_project_sha256=sha256(vba_project).hexdigest(),
    )


def validate_template(path: Path) -> WorkbookStructureSnapshot:
    return _structure_snapshot(path)


def fill_template(path: Path, rows: Sequence[OfficialPaymentRow]) -> bytes:
    if len(rows) > 500:
        raise InvalidOfficialPaymentWorkbookError(
            "official payment workbook accepts at most 500 rows"
        )
    expected_structure = validate_template(path)
    original_vba_project = _vba_project_bytes(path)
    workbook = load_workbook(path, keep_vba=True, data_only=False)
    upload_sheet = workbook[UPLOAD_SHEET]

    for row_number, row in enumerate(rows, start=2):
        for column_number, value in enumerate(row.as_cells(), start=1):
            upload_sheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )

    output = BytesIO()
    workbook.save(output)
    rendered = output.getvalue()
    rendered_vba_project = _vba_project_bytes(BytesIO(rendered))
    if rendered_vba_project != original_vba_project:
        raise InvalidOfficialPaymentWorkbookError(
            "VBA project bytes changed while filling the template"
        )

    if expected_structure.vba_project_sha256 != sha256(rendered_vba_project).hexdigest():
        raise InvalidOfficialPaymentWorkbookError(
            "VBA project identity changed while filling the template"
        )
    rendered_structure = _structure_snapshot(BytesIO(rendered))
    if rendered_structure != expected_structure:
        raise InvalidOfficialPaymentWorkbookError(
            "workbook structure changed while filling the template"
        )
    return rendered
